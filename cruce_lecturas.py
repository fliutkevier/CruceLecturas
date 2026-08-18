#!/usr/bin/env python3
"""
Cruce de Lecturas - Auditoría de medidores
==========================================
Toma la bajada del visualizador, descarta las pólizas ya cargadas en el excel
de errores, filtra por verificador y exporta los hallazgos nuevos ya
transformados al formato del excel de errores.

Los datos se copian VERBATIM desde la bajada (valor y formato de celda).
Solo se modifican: VERIFICADO (pasa a contener el auditor), la observación
(se le quitan fecha e iniciales) y se agrega una columna de fecha.
"""

import json
import os
import re
import subprocess
import sys
import tempfile
import threading
from datetime import date, datetime, time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

VERSION = "2.3"

URL_EXPORT = "https://psm.emaservicios.com.ar/visualizador/export"

# --------------------------------------------------------------------------
# Configuración
# --------------------------------------------------------------------------
COL_POLIZA = "POLIZA"
COL_TURNO = "TURNO"
COL_VERIFICADO = "VERIFICADO"
COL_OBS = "OBSERVACION DE AUDITORÍA"
COL_AUDITOR = "USUARIO AUDITOR"

COL_CATEGORIA = "CATEGORIA_ERROR"
COL_OBS = "OBSERVACION DE AUDITORÍA"
COL_AUDITOR = "USUARIO AUDITOR"
COL_POLIZA = "POLIZA"
COL_TURNO = "TURNO"

# Errores conocidos: separan el error del texto extra y permiten deducir la
# CATEGORIA cuando el visualizador no la trae cargada.
ERRORES_POR_DEFECTO = [
    "LECTURA ANTERIOR INCORRECTA",
    "LECTURA INCORRECTA",
    "ANOMALIA INCORRECTA",
    "FOTO MAL SACADA VISIBLE",
    "FOTO MAL SACADA",
    "MEDIDOR INCORRECTO",
]
ERRORES_CON_CODIGO_POR_DEFECTO = ["FALTA AVISO"]

# Categorias validas. La CATEGORIA sale de CATEGORIA_ERROR del visualizador;
# el detalle de la observacion va aparte, en DETALLE DE ERROR.
CATEGORIAS = [
    "LECTURA ANTERIOR INCORRECTA",
    "LECTURA INCORRECTA",
    "MEDIDOR INCORRECTO",
    "FOTO MAL SACADA VISIBLE",
    "FOTO MAL SACADA",
    "ANOMALIA INCORRECTA",
]
# Como lo escribe hoy el visualizador -> como lo queremos guardar.
ALIAS_CATEGORIA = {"ERROR DE ANOMALIA": "ANOMALIA INCORRECTA"}

ESTADO_INICIAL = "EN TRATAMIENTO"
RE_CODIGO = re.compile(r"[A-Z0-9]{1,3}$")

# <fecha> seguida de 2-3 iniciales en mayuscula, en cualquier parte del texto.
RE_FECHA_INICIALES = re.compile(
    r"\s*\b(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?\s+([A-ZÁÉÍÓÚÑ]{2,3})\b")
# <fecha> al comienzo del texto, sin iniciales detras.
RE_FECHA_INICIO = re.compile(r"^\s*(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?\s+")

FORMATO_FECHA = "DD/MM/YYYY"
FORMATO_HORA = "HH:MM"
HOJA_SALIDA = "REGISTRO"
HOJA_DESCONOCIDOS = "DESCONOCIDOS"
COL_MOTIVO = "MOTIVO"

# --- Formato unificado de salida (41 columnas) -----------------------------
# (columna final, campo del VISUALIZADOR, campo de la MUESTRA)
# Los tokens que empiezan con "@" los calcula la app.
ORDEN_SALIDA = [
    ("ORIGEN",                   "@ORIGEN",               "@ORIGEN"),
    ("TURNO",                    "TURNO",                 "TURNO"),
    ("BIMESTRE",                 "PERIODO",               "BIMESTRE"),
    ("AÑO",                      "AÑO DE FAC",            "ANIO"),
    ("RUTA",                     "RUTA",                  None),
    ("POLIZA",                   "POLIZA",                "POLIZA"),
    ("CLIENTE",                  "CLIENTE",               None),
    ("EMPRESA",                  None,                    "EMPRESA"),
    ("NOMBREFOTO",               None,                    "NOMBREFOTO"),
    ("AVISO",                    None,                    "AVISO"),
    ("SUBAVISOS",                None,                    "SUBAVISOS"),
    ("%AUDITORIA",               None,                    "%AUDITORIA"),
    ("IND-RECLAMO",              None,                    "IND-RECLAMO"),
    ("LECTURA CORREGIDA",        None,                    "LECTURA CORREGIDA"),
    ("FECHA LECTURA",            "@FECHA_VIS",            "@FECHA_NAT"),
    ("HORA LECTURA",             "@HORA_VIS",             "@HORA_NAT"),
    ("USUARIO",                  "@USUARIO",              "LECTURISTA"),
    ("LEGAJO",                   "@LEGAJO",               None),
    ("NOMBRE Y APELLIDO",        "@NOMBRE",               None),
    ("LOCALIDAD",                "LOCALIDAD",             "LOCALIDAD"),
    ("CALLE",                    "CALLE",                 "@CALLE_NAT"),
    ("ENTRE CALLES",             "ENTRE CALLES",          None),
    ("ACCESO PM",                "ACCESO PM",             None),
    ("DATOS DOMICILIO",          "DATOS_DOM_SERVICIO",    "@DOMICILIO_NAT"),
    ("N° DE MEDIDOR",            "@MEDIDOR",              "NRO-MEDIDOR"),
    ("AVISO-LECTOR",             "ORDENATIVO",            "AVISO-LECTOR"),
    ("COMENTARIO LECTOR",        "COMENTARIO LECTURISTA", None),
    ("LINK DE FOTO",             "FOTO 1",                None),
    ("LECTURA ANTERIOR",         "LEC ANTERIOR",          None),
    ("LECTURA ACTUAL",           "LEC ACTUAL",            "LECTURA-ACTUAL"),
    ("CONSUMO",                  "CONSUMO",               None),
    ("VERIFICADOR DE INICIO",    "USUARIO AUDITOR",       None),
    ("FECHA VERIFICADO",         "@FECHA_VERIF",          None),
    ("DETALLE DE ERROR",         "@DETALLE",              None),
    ("CATEGORIA",                "@CATEGORIA",            "@CATEGORIA_NAT"),
    ("AUDITORIA",                None,                    None),
    ("RESOLUCION",               None,                    None),
    ("DETALLE RESOLUCION",       None,                    None),
    ("FECHA DE AVISO AL LECTOR", None,                    None),
    ("VERIFICADOR DE CIERRE",    None,                    None),
    ("OBSERVACIONES",            None,                    None),
    ("ESTADO",                   "@ESTADO",               None),
]

# Se guardan como numero para que las dos fuentes convivan: la muestra manda
# todo como texto y el visualizador como numero.
COLS_NUMERICAS = {"TURNO", "BIMESTRE", "AÑO", "POLIZA", "USUARIO", "LEGAJO",
                  "N° DE MEDIDOR", "LECTURA ANTERIOR", "LECTURA ACTUAL", "CONSUMO"}

COLS_REQUERIDAS_VIS = [COL_POLIZA, COL_TURNO, COL_OBS, COL_AUDITOR]
COLS_REQUERIDAS_NAT = ["POLIZA", "TURNO", "ANIO", "BIMESTRE", "CORRECTO"]


VERIFICADORES_POR_DEFECTO = [
    {"nombre": "Tomas barragan", "iniciales": "TB"},
    {"nombre": "Juan Piccioli", "iniciales": "JP"},
    {"nombre": "Lucas Rodriguez", "iniciales": "LR"},
    {"nombre": "Candela Bolzan", "iniciales": "CB"},
    {"nombre": "Franco fliutkevier", "iniciales": "FL"},
    {"nombre": "Federico weber", "iniciales": "FW"},
]

URL_EXPORT = "https://psm.emaservicios.com.ar/visualizador/export"


def recurso(nombre):
    """Ruta a un archivo empaquetado, tanto en desarrollo como compilado."""
    base = getattr(sys, "_MEIPASS", None) or os.path.dirname(
        sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__))
    return os.path.join(base, nombre)


def abrir_archivo(ruta):
    """Abre el archivo con la aplicación por defecto del sistema."""
    if not os.path.isfile(ruta):
        return False
    try:
        if sys.platform == "win32":
            os.startfile(ruta)                                  # noqa: S606
        elif sys.platform == "darwin":
            return subprocess.run(["open", ruta], check=False).returncode == 0
        else:
            return subprocess.run(["xdg-open", ruta], check=False).returncode == 0
        return True
    except Exception:
        return False


def descargar_turno(turno, periodo, anio, carpeta=None, timeout=180):
    """Descarga el export del visualizador. Devuelve la ruta del archivo."""
    try:
        import requests
    except ImportError:
        raise RuntimeError(
            "Falta la librería 'requests'.\nInstalala con: pip install requests")

    params = {
        "turno": turno, "periodo": periodo, "anio": anio,
        "fecha_desde": "", "fecha_hasta": "",
        "filtro_nro_medidor": "", "filtro_poliza": "",
        "origen": "lecturas",
    }
    try:
        r = requests.get(URL_EXPORT, params=params, timeout=timeout)
        r.raise_for_status()
    except Exception as exc:
        raise RuntimeError(f"No se pudo descargar el turno.\n\n{exc}")

    ctype = (r.headers.get("Content-Type") or "").lower()
    if not any(t in ctype for t in ("spreadsheet", "excel", "octet-stream")):
        raise RuntimeError(
            "El servidor no devolvió un Excel.\n\n"
            f"Content-Type: {ctype or '(vacío)'}\n"
            "Verificá que el turno, período y año existan.")
    if len(r.content) < 1024:
        raise RuntimeError("El archivo descargado está vacío o es demasiado chico.")
    if not r.content.startswith(b"PK"):
        raise RuntimeError("El archivo descargado no es un .xlsx válido.")

    carpeta = carpeta or carpeta_descargas()
    os.makedirs(carpeta, exist_ok=True)
    sello = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta = os.path.join(
        carpeta, f"lecturas_turno{turno}_periodo{periodo}_anio{anio}_{sello}.xlsx")
    with open(ruta, "wb") as f:
        f.write(r.content)
    return ruta


def _carpeta_app():
    return os.path.dirname(
        sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__))


def _carpeta_usuario():
    carpeta = os.path.join(os.environ.get("APPDATA", os.path.expanduser("~")),
                           "CruceLecturas")
    os.makedirs(carpeta, exist_ok=True)
    return carpeta


ARCHIVO_REDIRECCION = "config_ruta.txt"


def _ruta_redirigida():
    """Lee config_ruta.txt, que apunta a una config compartida en otro lado.

    Permite que cada PC tenga su copia local de la app pero todas lean la misma
    configuración. El archivo contiene una sola línea con la ruta completa.
    """
    base = _carpeta_app()
    for carpeta in (base, os.path.dirname(base)):
        txt = os.path.join(carpeta, ARCHIVO_REDIRECCION)
        if not os.path.isfile(txt):
            continue
        try:
            with open(txt, encoding="utf-8-sig") as f:
                for linea in f:
                    linea = linea.strip().strip('"')
                    if linea and not linea.startswith("#"):
                        return linea
        except OSError:
            pass
    return None


def ruta_cache_config():
    """Copia local de la última config compartida que se pudo leer."""
    return os.path.join(_carpeta_usuario(), "config_cache.json")


def ruta_config():
    r"""Config COMPARTIDA: verificadores y errores conocidos.

    Se busca primero en la carpeta PADRE de la aplicación, para que actualizar
    la app (reemplazar su carpeta entera) no se lleve puesta la configuración:

        \\servidor\Lecturas\CruceLecturas\
            config.json          <- sobrevive a las actualizaciones
            App\
                CruceLecturas.exe
                data\

    Si no hay carpeta padre escribible, cae junto al ejecutable y, como último
    recurso, al perfil del usuario (ahí deja de ser compartida).
    """
    redirigida = _ruta_redirigida()
    if redirigida:
        return redirigida

    base = _carpeta_app()
    padre = os.path.dirname(base)
    candidatos = [os.path.join(padre, "config.json"),
                  os.path.join(base, "config.json")]
    # 1. Si ya existe en alguno de los dos lugares, ese manda.
    for ruta in candidatos:
        if os.path.isfile(ruta):
            return ruta
    # 2. Si no existe, se crea en el primero que se pueda escribir.
    for ruta in candidatos:
        carpeta = os.path.dirname(ruta)
        if carpeta and os.path.isdir(carpeta) and os.access(carpeta, os.W_OK):
            return ruta
    return os.path.join(_carpeta_usuario(), "config.json")


def ruta_config_local():
    """Config INDIVIDUAL: carpeta de descargas. Nunca se comparte."""
    return os.path.join(_carpeta_usuario(), "config_local.json")


def _leer_json(ruta):
    try:
        with open(ruta, encoding="utf-8") as f:
            datos = json.load(f)
        if isinstance(datos, dict):
            return datos
    except (OSError, ValueError):
        pass
    return {}


def _escribir_json(ruta, cambios):
    """Actualiza solo las claves indicadas, sin pisar el resto del archivo."""
    cfg = _leer_json(ruta)
    cfg.update(cambios)
    try:
        with open(ruta, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
        return True
    except OSError:
        return False


def cargar_config():
    """Lee la config compartida. Si no está disponible, usa la copia en caché.

    Así un corte de red no deja la app sin verificadores ni errores.
    """
    ruta = ruta_config()
    if os.path.isfile(ruta):
        datos = _leer_json(ruta)
        if datos:
            try:                                  # refrescar la copia local
                with open(ruta_cache_config(), "w", encoding="utf-8") as f:
                    json.dump(datos, f, ensure_ascii=False, indent=2)
            except OSError:
                pass
            return datos
    cache = ruta_cache_config()
    if os.path.isfile(cache):
        return _leer_json(cache)
    return {}


def config_disponible():
    """True si la config compartida se puede leer ahora mismo.

    Si ya hubo una config conocida (existe caché) y ahora no se puede leer,
    devuelve False: es un estado degradado que conviene mostrar.
    """
    ruta = ruta_config()
    if os.path.isfile(ruta):
        return True
    if os.path.isfile(ruta_cache_config()):
        return False          # antes se leía y ahora no
    carpeta = os.path.dirname(ruta)
    return bool(carpeta) and os.path.isdir(carpeta)   # primera ejecución


def guardar_config(cambios):
    return _escribir_json(ruta_config(), cambios)


def cargar_errores():
    cfg = cargar_config()
    simples = cfg.get("errores")
    codigos = cfg.get("errores_con_codigo")
    if not isinstance(simples, list) or not simples:
        simples = list(ERRORES_POR_DEFECTO)
    if not isinstance(codigos, list) or not codigos:
        codigos = list(ERRORES_CON_CODIGO_POR_DEFECTO)
    return simples, codigos


def guardar_errores(simples, codigos):
    return guardar_config({"errores": simples, "errores_con_codigo": codigos})


def cargar_verificadores():
    datos = cargar_config().get("verificadores")
    if isinstance(datos, list) and datos:
        return datos
    return [dict(v) for v in VERIFICADORES_POR_DEFECTO]


def guardar_verificadores(lista):
    return guardar_config({"verificadores": lista})


def carpeta_descargas_por_defecto():
    return os.path.join(tempfile.gettempdir(), "CruceLecturas")


def carpeta_descargas():
    """Carpeta configurada por ESTE usuario, o la temporal si no hay ninguna."""
    return (_leer_json(ruta_config_local()).get("carpeta_descargas")
            or carpeta_descargas_por_defecto())


def guardar_carpeta_descargas(ruta):
    return _escribir_json(ruta_config_local(), {"carpeta_descargas": ruta})


# --------------------------------------------------------------------------
# Núcleo
# --------------------------------------------------------------------------
def normalizar_poliza(valor):
    """Canoniza una póliza SOLO para comparar. Nunca se escribe al archivo."""
    if valor is None:
        return None
    if isinstance(valor, float):
        if valor != valor:
            return None
        if valor.is_integer():
            return str(int(valor))
    texto = str(valor).strip()
    if not texto or texto.lower() in ("nan", "none"):
        return None
    if re.fullmatch(r"\d+\.0+", texto):
        texto = texto.split(".")[0]
    return texto


def parsear_polizas_pegadas(texto):
    """Convierte el pegado crudo del excel de errores en un set de pólizas."""
    if not texto or not texto.strip():
        return set()
    polizas = set()
    for token in re.split(r"[\s,;]+", texto.strip()):
        norm = normalizar_poliza(token)
        if norm and norm.isdigit():
            polizas.add(norm)
    return polizas


def _armar_fecha(dia, mes, anio, hoy):
    """Reconstruye la fecha. Si no trae año, infiere el más cercano."""
    try:
        dia, mes = int(dia), int(mes)
        if anio:
            anio = int(anio)
            if anio < 100:
                anio += 2000
        else:
            anio = hoy.year
            # Observación de diciembre leída en enero: pertenece al año anterior.
            if mes - hoy.month > 6:
                anio -= 1
        return date(anio, mes, dia)
    except (ValueError, TypeError):
        return None


def limpiar_observacion(texto, hoy=None):
    """Quita fecha e iniciales del texto y devuelve (texto_limpio, fecha, n_autores).

    n_autores permite detectar cadenas con más de un auditor, que en operación
    normal no deberían llegar (ya estarían cargadas en el excel de errores).
    """
    hoy = hoy or date.today()
    if texto is None:
        return None, None, 0
    original = str(texto)
    fechas, iniciales = [], set()

    def _capturar(m):
        f = _armar_fecha(m.group(1), m.group(2), m.group(3), hoy)
        if f:
            fechas.append(f)
        iniciales.add(m.group(4))
        return " "

    limpio = RE_FECHA_INICIALES.sub(_capturar, original)

    if not fechas:  # formato con la fecha adelante y sin iniciales
        m = RE_FECHA_INICIO.match(limpio)
        if m:
            f = _armar_fecha(m.group(1), m.group(2), m.group(3), hoy)
            if f:
                fechas.append(f)
                limpio = limpio[m.end():]

    limpio = re.sub(r"\s{2,}", " ", limpio).strip(" +-")
    if not limpio:            # nunca dejar la observación vacía
        limpio = original.strip()
    fecha = max(fechas) if fechas else None
    return limpio, fecha, len(iniciales)


def separar_error_detalle(texto, errores=None, con_codigo=None):
    """Parte la observación en (error_conocido, detalle).

    Si no reconoce ningún error, devuelve (texto, "") sin tocar nada: nunca
    descarta información por no estar en la lista.
    """
    if texto is None:
        return None, None
    t = str(texto).strip()
    if not t:
        return t, ""
    if errores is None or con_codigo is None:
        cargados = cargar_errores()
        errores = errores if errores is not None else cargados[0]
        con_codigo = con_codigo if con_codigo is not None else cargados[1]

    tu = t.upper()

    # 1. Prefijos que absorben códigos de anomalía (FALTA AVISO G, 34, ZT...).
    for pref in sorted(con_codigo, key=len, reverse=True):
        if tu.startswith(pref.upper()):
            resto = t[len(pref):].strip()
            toks = resto.split()
            n = 0
            while n < len(toks) and RE_CODIGO.match(toks[n]):
                n += 1
            error = (pref + " " + " ".join(toks[:n])).strip()
            return error, " ".join(toks[n:]).strip()

    # 2. Errores literales. Del más largo al más corto, para que
    #    "FOTO MAL SACADA VISIBLE" gane sobre "FOTO MAL SACADA".
    for e in sorted(errores, key=len, reverse=True):
        if tu.startswith(e.upper()):
            return t[:len(e)].strip(), t[len(e):].strip()

    return t, ""


def _texto(valor):
    return "" if valor is None else str(valor).strip()


def _num(valor):
    """Convierte a numero si se puede. Sirve para unificar tipos entre fuentes:
    la muestra manda '858806' (texto) y el visualizador 858806 (numero)."""
    if valor is None:
        return None
    if isinstance(valor, bool):
        return valor
    if isinstance(valor, (int, float)):
        return int(valor) if float(valor).is_integer() else valor
    t = str(valor).strip()
    if not t:
        return None
    try:
        f = float(t)
        return int(f) if f.is_integer() else f
    except ValueError:
        return valor


def partir_lecturista(valor):
    """'31 - 8466 - ALEGRE, EZEQUIEL' -> ('31', '8466', 'ALEGRE, EZEQUIEL')."""
    t = _texto(valor)
    if not t:
        return None, None, None
    partes = [p.strip() for p in t.split(" - ", 2)]
    while len(partes) < 3:
        partes.append(None)
    return partes[0] or None, partes[1] or None, partes[2] or None


def fecha_naturgy(valor):
    """La muestra manda DDMMAA como numero, sin cero adelante: 30826 = 03/08/26."""
    t = re.sub(r"\D", "", _texto(valor))
    if not t or len(t) > 6:
        return None
    t = t.zfill(6)
    try:
        return date(2000 + int(t[4:6]), int(t[2:4]), int(t[0:2]))
    except ValueError:
        return None


def hora_naturgy(valor):
    """HHMM como numero, sin cero adelante: 901 = 09:01."""
    t = re.sub(r"\D", "", _texto(valor))
    if not t or len(t) > 4:
        return None
    t = t.zfill(4)
    try:
        return time(int(t[0:2]), int(t[2:4]))
    except ValueError:
        return None


def fecha_visualizador(valor):
    """El visualizador manda 'DD-MM-AA' (o ya una fecha real)."""
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    t = _texto(valor)
    for sep in ("-", "/"):
        partes = t.split(sep)
        if len(partes) == 3 and all(p.strip().isdigit() for p in partes):
            d, m, a = (int(p) for p in partes)
            if a < 100:
                a += 2000
            try:
                return date(a, m, d)
            except ValueError:
                return None
    return None


def hora_visualizador(valor):
    """El visualizador manda 'HH:MM' (o ya una hora real)."""
    if isinstance(valor, datetime):
        return valor.time()
    if isinstance(valor, time):
        return valor
    t = _texto(valor)
    partes = t.split(":")
    if len(partes) >= 2 and all(p.strip().isdigit() for p in partes[:2]):
        try:
            return time(int(partes[0]), int(partes[1]))
        except ValueError:
            return None
    return None


def domicilio_naturgy(portal, escalera, piso, puerta):
    """PORTAL/ESCALERA/PISO/PUERTA -> 'P:PB-PU:01', igual que el visualizador."""
    partes = []
    for etiqueta, valor in (("PT", portal), ("E", escalera),
                            ("P", piso), ("PU", puerta)):
        t = _texto(valor)
        if t:
            partes.append(f"{etiqueta}:{t}")
    return "-".join(partes) or None


def calle_naturgy(calle, finca):
    """La muestra separa la altura en FINCA; el visualizador ya la trae junta."""
    return " ".join(p for p in (_texto(calle), _texto(finca)) if p) or None


def normalizar_categoria(valor):
    """Pasa a mayusculas y aplica los alias (ERROR DE ANOMALIA -> ANOMALIA INCORRECTA)."""
    t = _texto(valor).upper()
    if not t:
        return None
    return ALIAS_CATEGORIA.get(t, t)


def limpiar_medidor(valor):
    """El visualizador manda '632420.0'."""
    t = _texto(valor)
    if re.fullmatch(r"\d+\.0+", t):
        t = t.split(".")[0]
    return _num(t) if t else None


def _valor_final(columna, valor):
    """Aplica la unificacion de tipos segun la columna."""
    if valor is None or valor == "":
        return None
    if columna in COLS_NUMERICAS:
        return _num(valor)
    return valor


def _formato(columna, valor):
    if isinstance(valor, date) and not isinstance(valor, datetime):
        return FORMATO_FECHA
    if isinstance(valor, time):
        return FORMATO_HORA
    return "General"


def _armar_fila(calculados, fila, idx, indice_fuente):
    """Arma una fila de salida en el orden de ORDEN_SALIDA.

    indice_fuente: 1 = visualizador, 2 = muestra.
    """
    salida = []
    for entrada in ORDEN_SALIDA:
        col = entrada[0]
        campo = entrada[indice_fuente]
        if campo is None:
            valor = None
        elif campo.startswith("@"):
            valor = calculados.get(campo)
        else:
            pos = idx.get(campo)
            valor = fila[pos].value if pos is not None else None
        valor = _valor_final(col, valor)
        salida.append((valor, _formato(col, valor)))
    return salida


def _abrir(ruta):
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows()
    try:
        cabecera = next(it)
    except StopIteration:
        wb.close()
        raise ValueError("El archivo está vacío.")
    headers = [_texto(c.value) for c in cabecera]
    return wb, it, headers


def headers_salida():
    return [c[0] for c in ORDEN_SALIDA]


# --------------------------------------------------------------------------
# Procesamiento: VISUALIZADOR (cruce)
# --------------------------------------------------------------------------
def procesar_cruce(ruta_bajada, texto_polizas, verificadores=None, hoy=None,
                   errores=None, con_codigo=None):
    """Cruce del visualizador. Devuelve (headers, filas, stats)."""
    hoy = hoy or date.today()
    if errores is None or con_codigo is None:
        errores, con_codigo = cargar_errores()

    wb, it, headers = _abrir(ruta_bajada)
    faltantes = [c for c in COLS_REQUERIDAS_VIS if c not in headers]
    if faltantes:
        wb.close()
        raise ValueError("La bajada no tiene las columnas esperadas: " + ", ".join(faltantes))
    idx = {h: i for i, h in enumerate(headers)}

    seleccion = {v.strip().lower() for v in (verificadores or []) if v and v.strip()}
    excluidas = parsear_polizas_pegadas(texto_polizas)

    filas_salida, vistas, desconocidos, sin_reconocer = [], set(), set(), []
    apartadas = []

    def _apartar(fila, idx, motivo):
        """Guarda la fila cruda (sin transformar) para no perder nada."""
        apartadas.append([motivo] + [c.value for c in fila])
    total_filas = total_con_obs = ya_cargados = duplicados = 0
    filtrados = limpiadas = multi_autor = sin_fecha = sin_categoria = 0
    turno = ""

    for fila in it:
        if all(c.value is None for c in fila):
            continue
        total_filas += 1
        if not turno:
            turno = normalizar_poliza(fila[idx[COL_TURNO]].value) or ""

        if not _texto(fila[idx[COL_OBS]].value):
            continue
        total_con_obs += 1

        poliza = normalizar_poliza(fila[idx[COL_POLIZA]].value)
        if poliza in excluidas:
            ya_cargados += 1
            continue

        auditor = fila[idx[COL_AUDITOR]].value
        nombre = _texto(auditor)
        if seleccion and nombre.lower() not in seleccion:
            if nombre:
                desconocidos.add(nombre)
            filtrados += 1
            _apartar(fila, idx, f"auditor no seleccionado: {nombre or '(vacío)'}")
            continue

        if poliza in vistas:
            duplicados += 1
            _apartar(fila, idx, "póliza repetida dentro de la misma bajada")
            continue
        vistas.add(poliza)

        obs_cruda = fila[idx[COL_OBS]].value
        obs_limpia, fecha_obs, n_aut = limpiar_observacion(obs_cruda, hoy)
        if obs_limpia != _texto(obs_cruda):
            limpiadas += 1
        if n_aut > 1:
            multi_autor += 1
        if fecha_obs is None:
            fecha_obs = hoy
            sin_fecha += 1

        # CATEGORIA: la del visualizador si viene; si no, se deduce del texto.
        # CATEGORIA: sale del desplegable del visualizador, no del texto.
        cat = normalizar_categoria(
            fila[idx[COL_CATEGORIA]].value if COL_CATEGORIA in idx else None)
        if not cat:
            sin_categoria += 1
            sin_reconocer.append(obs_limpia)   # queda con CATEGORIA vacia, no se aparta

        # DETALLE: la observacion limpia. Si arranca repitiendo una categoria
        # (dato viejo, antes de que existiera el desplegable), se le quita para
        # no duplicar el mismo texto en dos columnas. "FALTA AVISO ..." NO se
        # toca: es un detalle, no una categoria.
        detalle = obs_limpia
        if cat and detalle.upper().startswith(cat.upper()):
            # Solo se quita si REPITE la categoria ya guardada. Si el texto dice
            # otra cosa que el desplegable, se conserva: esa discrepancia es
            # informacion, no ruido.
            detalle = detalle[len(cat):].strip(" +-")

        usuario, legajo, nombre_ap = partir_lecturista(
            fila[idx["NUM LECURISTA"]].value if "NUM LECURISTA" in idx else None)

        calc = {
            "@ORIGEN": "INTERNO",
            "@FECHA_VIS": fecha_visualizador(fila[idx["FECHA"]].value) if "FECHA" in idx else None,
            "@HORA_VIS": hora_visualizador(fila[idx["HORA"]].value) if "HORA" in idx else None,
            "@USUARIO": usuario, "@LEGAJO": legajo, "@NOMBRE": nombre_ap,
            "@MEDIDOR": limpiar_medidor(fila[idx["NRO MEDIDOR"]].value) if "NRO MEDIDOR" in idx else None,
            "@FECHA_VERIF": fecha_obs,
            "@DETALLE": detalle or None,
            "@CATEGORIA": cat,
            "@ESTADO": ESTADO_INICIAL,
        }
        filas_salida.append(_armar_fila(calc, fila, idx, 1))

    wb.close()
    stats = {
        "fuente": "visualizador",
        "total_filas": total_filas, "total_con_obs": total_con_obs,
        "excluidas_pegadas": len(excluidas), "ya_cargados": ya_cargados,
        "filtrados": filtrados, "duplicados": duplicados, "limpiadas": limpiadas,
        "multi_autor": multi_autor, "sin_fecha": sin_fecha,
        "sin_categoria": sin_categoria,
        "desconocidos": sorted(desconocidos),
        "sin_reconocer": sorted(set(sin_reconocer)),
        "columnas_ausentes": [],
        "nuevos": len(filas_salida), "turno": turno,
        "primer_cruce": len(excluidas) == 0,
        "apartadas": len(apartadas),
        "headers_origen": [COL_MOTIVO] + headers,
    }
    return headers_salida(), filas_salida, stats, apartadas


# --------------------------------------------------------------------------
# Procesamiento: MUESTRA DE NATURGY
# --------------------------------------------------------------------------
def procesar_muestra(ruta_muestra):
    """Convierte la muestra de Naturgy al formato unificado.

    No filtra ni deduplica: entran todas las filas, correctas e incorrectas.
    """
    wb, it, headers = _abrir(ruta_muestra)
    faltantes = [c for c in COLS_REQUERIDAS_NAT if c not in headers]
    if faltantes:
        wb.close()
        raise ValueError("La muestra no tiene las columnas esperadas: " + ", ".join(faltantes))
    idx = {h: i for i, h in enumerate(headers)}

    def v(fila, campo):
        return fila[idx[campo]].value if campo in idx else None

    filas_salida = []
    total = sin_fecha = sin_hora = 0
    turno = ""
    categorias = {}

    for fila in it:
        if all(c.value is None for c in fila):
            continue
        total += 1
        if not turno:
            turno = normalizar_poliza(v(fila, "TURNO")) or ""

        f = fecha_naturgy(v(fila, "FECHA-LECT"))
        h = hora_naturgy(v(fila, "HORA-LECT"))
        if f is None:
            sin_fecha += 1
        if h is None:
            sin_hora += 1

        cat = _texto(v(fila, "CORRECTO")).upper() or None
        if cat:
            categorias[cat] = categorias.get(cat, 0) + 1

        calc = {
            "@ORIGEN": "CLIENTE",
            "@FECHA_NAT": f, "@HORA_NAT": h,
            "@CALLE_NAT": calle_naturgy(v(fila, "CALLE"), v(fila, "FINCA")),
            "@DOMICILIO_NAT": domicilio_naturgy(v(fila, "PORTAL"), v(fila, "ESCALERA"),
                                                v(fila, "PISO"), v(fila, "PUERTA")),
            "@CATEGORIA_NAT": cat,
        }
        filas_salida.append(_armar_fila(calc, fila, idx, 2))

    wb.close()
    stats = {
        "fuente": "muestra", "total_filas": total, "nuevos": len(filas_salida),
        "sin_fecha": sin_fecha, "sin_hora": sin_hora,
        "categorias": categorias, "turno": turno,
        "columnas_ausentes": [c for c in ("EMPRESA", "NOMBREFOTO", "AVISO", "SUBAVISOS",
                                          "%AUDITORIA", "IND-RECLAMO", "LECTURA CORREGIDA",
                                          "LECTURA-ACTUAL", "AVISO-LECTOR", "LOCALIDAD",
                                          "NRO-MEDIDOR", "LECTURISTA")
                              if c not in idx],
    }
    return headers_salida(), filas_salida, stats


def exportar(headers, filas, turno, carpeta_destino, prefijo="Cruce",
             apartadas=None, headers_apartadas=None):
    """Escribe el xlsx en el formato unificado (hoja REGISTRO).

    Si hay filas apartadas, se agregan en una segunda hoja DESCONOCIDOS con
    su contenido ORIGINAL mas una columna MOTIVO. Nada se descarta en silencio.
    """
    sello = datetime.now().strftime("%Y-%m-%d_%H-%M")
    ruta = os.path.join(carpeta_destino, f"{prefijo}Turno{turno or 'SD'}_{sello}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = HOJA_SALIDA

    fuente = Font(name="Arial", size=10)
    ws.append(["" if h is None else h for h in headers])
    for celda in ws[1]:
        celda.font = Font(name="Arial", size=10, bold=True)
        celda.alignment = Alignment(vertical="center", wrap_text=True)

    for fila in filas:
        ws.append([v for v, _ in fila])
        r = ws.max_row
        for c, (_, fmt) in enumerate(fila, start=1):
            celda = ws.cell(row=r, column=c)
            celda.font = fuente
            if fmt:
                celda.number_format = fmt

    ws.freeze_panes = "A2"
    anchos = [len(str(h or "")) for h in headers]
    for fila in filas:
        for i, (v, _) in enumerate(fila):
            if v is not None:
                anchos[i] = max(anchos[i], len(str(v)))
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(ancho + 2, 10), 45)

    if apartadas:
        ws2 = wb.create_sheet(HOJA_DESCONOCIDOS)
        ws2.append([h if h is not None else "" for h in (headers_apartadas or [])])
        for celda in ws2[1]:
            celda.font = Font(name="Arial", size=10, bold=True)
            celda.alignment = Alignment(vertical="center", wrap_text=True)
        ws2["A1"].fill = PatternFill("solid", fgColor="FFF2CC")
        for fila in apartadas:
            ws2.append(fila)
            for celda in ws2[ws2.max_row]:
                celda.font = fuente
        ws2.freeze_panes = "B2"
        anchos2 = [len(str(h or "")) for h in (headers_apartadas or [])]
        for fila in apartadas:
            for i, v in enumerate(fila):
                if i < len(anchos2) and v is not None:
                    anchos2[i] = max(anchos2[i], len(str(v)))
        for i, ancho in enumerate(anchos2, start=1):
            ws2.column_dimensions[get_column_letter(i)].width = min(max(ancho + 2, 10), 50)

    wb.save(ruta)
    return ruta


# --------------------------------------------------------------------------
# Interfaz gráfica
# --------------------------------------------------------------------------
def lanzar_ui():
    import tkinter as tk
    from tkinter import filedialog, messagebox, simpledialog, ttk

    estado = {"bajada": None, "muestra": None,
              "verificadores": cargar_verificadores(), "vars": {}}

    if sys.platform == "win32":
        try:
            import ctypes
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(
                f"Zelkov.CruceLecturas.{VERSION}")
        except Exception:
            pass

    root = tk.Tk()
    root.title(f"Cruce de Lecturas v{VERSION}")
    root.geometry("720x880")
    root.minsize(660, 720)

    for candidato in ("icono.ico", "icon.ico"):
        try:
            ruta_ico = recurso(candidato)
            if os.path.exists(ruta_ico):
                root.iconbitmap(default=ruta_ico)
                break
        except Exception:
            pass

    cont = ttk.Frame(root, padding=14)
    cont.pack(fill="both", expand=True)
    ttk.Label(cont, text="Cruce de Lecturas",
              font=("Segoe UI", 15, "bold")).pack(anchor="w")
    ttk.Label(cont, foreground="#555", wraplength=660, justify="left",
              text="Las dos fuentes terminan en el mismo formato, listo para "
                   "pegar en el archivo de seguimiento.").pack(anchor="w", pady=(2, 10))

    nb = ttk.Notebook(cont)
    nb.pack(fill="both", expand=True)
    tab_vis = ttk.Frame(nb, padding=12)
    tab_nat = ttk.Frame(nb, padding=12)
    nb.add(tab_vis, text="  Visualizador (verificadores)  ")
    nb.add(tab_nat, text="  Muestra de Naturgy  ")

    def guardar_resultado(headers, filas, turno, prefijo, detalle,
                          apartadas=None, headers_apartadas=None):
        destino = filedialog.askdirectory(title="¿Dónde guardo el resultado?")
        if not destino:
            return
        try:
            ruta = exportar(headers, filas, turno, destino, prefijo,
                            apartadas, headers_apartadas)
        except PermissionError:
            messagebox.showerror("Archivo bloqueado",
                                 "No se pudo escribir. ¿Está abierto en Excel?")
            return
        detalle.append(f"\nArchivo: {os.path.basename(ruta)}")
        messagebox.showinfo("Listo", "\n".join(detalle))
        if not abrir_archivo(ruta):
            messagebox.showinfo("Archivo generado",
                                f"No se pudo abrir automáticamente.\n\nEstá en:\n{ruta}")

    # =====================================================================
    # PESTAÑA 1 — VISUALIZADOR
    # =====================================================================
    ttk.Label(tab_vis, text="1 · Bajada del visualizador",
              font=("Segoe UI", 10, "bold")).pack(anchor="w")
    lbl_archivo = ttk.Label(tab_vis, text="Ningún archivo seleccionado", foreground="#888")

    def fijar_bajada(ruta):
        estado["bajada"] = ruta
        lbl_archivo.config(text=os.path.basename(ruta), foreground="#000")

    imp = ttk.Frame(tab_vis)
    imp.pack(fill="x", pady=(4, 2))
    ttk.Label(imp, text="Turno").pack(side="left")
    e_turno = ttk.Entry(imp, width=6); e_turno.pack(side="left", padx=(4, 10))
    ttk.Label(imp, text="Período").pack(side="left")
    e_per = ttk.Entry(imp, width=6); e_per.pack(side="left", padx=(4, 10))
    ttk.Label(imp, text="Año").pack(side="left")
    e_anio = ttk.Entry(imp, width=8); e_anio.insert(0, str(date.today().year))
    e_anio.pack(side="left", padx=(4, 10))
    btn_imp = ttk.Button(imp, text="Importar turno"); btn_imp.pack(side="left")

    fc = ttk.Frame(tab_vis); fc.pack(fill="x", pady=(2, 2))
    ttk.Label(fc, text="Guardar bajadas en:", foreground="#555").pack(side="left")
    lbl_carpeta = ttk.Label(fc, foreground="#333"); lbl_carpeta.pack(side="left", padx=(6, 6))

    def refrescar_carpeta():
        ruta = carpeta_descargas()
        existe = os.path.isdir(ruta) or ruta == carpeta_descargas_por_defecto()
        texto = ruta if len(ruta) <= 48 else "..." + ruta[-45:]
        if not existe:
            texto += "   (no disponible)"
        lbl_carpeta.config(text=texto, foreground="#333" if existe else "#B00")

    def cambiar_carpeta():
        elegida = filedialog.askdirectory(title="¿Dónde guardo las bajadas?")
        if not elegida:
            return
        if not guardar_carpeta_descargas(elegida):
            messagebox.showwarning("Sin permisos",
                                   "Se usa ahora, pero no se pudo guardar la preferencia.")
        refrescar_carpeta()

    ttk.Button(fc, text="Cambiar...", command=cambiar_carpeta).pack(side="left")
    refrescar_carpeta()

    f1 = ttk.Frame(tab_vis); f1.pack(fill="x", pady=(2, 4))
    ttk.Label(f1, text="Alternativa:", foreground="#888").pack(side="left", padx=(0, 6))

    def elegir():
        ruta = filedialog.askopenfilename(
            title="Seleccioná la bajada del visualizador",
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")])
        if ruta:
            fijar_bajada(ruta)

    btn_elegir = ttk.Button(f1, text="Seleccionar archivo...", command=elegir)
    btn_elegir.pack(side="left")
    lbl_archivo.pack(anchor="w", pady=(0, 10))

    def importar():
        turno, per, anio = (e_turno.get().strip(), e_per.get().strip(), e_anio.get().strip())
        if not (turno.isdigit() and per.isdigit() and anio.isdigit()):
            messagebox.showwarning("Datos incompletos",
                                   "Turno, período y año tienen que ser números.")
            return
        destino = carpeta_descargas()
        if not os.path.isdir(destino):
            try:
                os.makedirs(destino, exist_ok=True)
            except OSError:
                if not messagebox.askyesno("Carpeta no disponible",
                        f"No se puede acceder a:\n{destino}\n\n¿Usar la temporal?"):
                    return
                destino = carpeta_descargas_por_defecto()
        btn_imp.config(state="disabled", text="Descargando...")
        btn_elegir.config(state="disabled")
        lbl_archivo.config(text="Descargando del visualizador...", foreground="#888")

        def trabajo():
            try:
                ruta = descargar_turno(turno, per, anio, destino)
            except Exception as exc:
                # Se captura por valor: Python borra `exc` al salir del except,
                # y el lambda corre despues (fallaria con NameError).
                mensaje = str(exc)
                root.after(0, lambda m=mensaje: fin(None, m))
            else:
                root.after(0, lambda r=ruta: fin(r, None))

        def fin(ruta, error):
            btn_imp.config(state="normal", text="Importar turno")
            btn_elegir.config(state="normal")
            if error:
                lbl_archivo.config(text="Ningún archivo seleccionado", foreground="#888")
                messagebox.showerror("Error al importar", error)
            else:
                fijar_bajada(ruta); refrescar_carpeta()

        threading.Thread(target=trabajo, daemon=True).start()

    btn_imp.config(command=importar)

    ttk.Label(tab_vis, text="2 · Verificadores a incluir",
              font=("Segoe UI", 10, "bold")).pack(anchor="w")
    ttk.Label(tab_vis, foreground="#555", wraplength=640, justify="left",
              text="Si no tildás ninguno, entran todos.").pack(anchor="w", pady=(2, 4))
    caja = ttk.Frame(tab_vis, relief="solid", borderwidth=1); caja.pack(fill="x")
    lista = ttk.Frame(caja, padding=6); lista.pack(fill="x")

    def redibujar():
        for w in lista.winfo_children():
            w.destroy()
        estado["vars"].clear()
        for i, v in enumerate(estado["verificadores"]):
            var = tk.BooleanVar(value=False)
            estado["vars"][v["nombre"]] = var
            ttk.Checkbutton(lista, variable=var,
                            text=f'{v["nombre"]}  ({v.get("iniciales","--")})'
                            ).grid(row=i // 2, column=i % 2, sticky="w", padx=6, pady=1)

    def agregar():
        nombre = simpledialog.askstring("Agregar verificador",
            "Nombre exacto como figura en USUARIO AUDITOR:", parent=root)
        if not nombre or not nombre.strip():
            return
        nombre = nombre.strip()
        if any(v["nombre"].lower() == nombre.lower() for v in estado["verificadores"]):
            messagebox.showinfo("Ya existe", "Ese verificador ya está en la lista.")
            return
        ini = simpledialog.askstring("Agregar verificador", "Iniciales (opcional):",
                                     parent=root) or ""
        estado["verificadores"].append({"nombre": nombre, "iniciales": ini.strip().upper()})
        estado["verificadores"].sort(key=lambda v: v["nombre"].lower())
        if not guardar_verificadores(estado["verificadores"]):
            messagebox.showwarning("Sin permisos", "Se agregó, pero no se pudo guardar.")
        redibujar()

    def quitar():
        tildados = [n for n, v in estado["vars"].items() if v.get()]
        if not tildados:
            messagebox.showinfo("Quitar", "Tildá los que querés quitar.")
            return
        if not messagebox.askyesno("Confirmar", "¿Quitar?\n\n" + "\n".join(tildados)):
            return
        estado["verificadores"] = [v for v in estado["verificadores"]
                                   if v["nombre"] not in tildados]
        guardar_verificadores(estado["verificadores"])
        redibujar()

    f2 = ttk.Frame(tab_vis); f2.pack(fill="x", pady=(4, 10))
    ttk.Button(f2, text="Agregar...", command=agregar).pack(side="left")
    ttk.Button(f2, text="Quitar tildados", command=quitar).pack(side="left", padx=6)
    ttk.Button(f2, text="Tildar todos",
               command=lambda: [v.set(True) for v in estado["vars"].values()]).pack(side="left")
    ttk.Button(f2, text="Destildar todos",
               command=lambda: [v.set(False) for v in estado["vars"].values()]
               ).pack(side="left", padx=6)
    redibujar()

    ttk.Label(tab_vis, text="3 · Pólizas ya cargadas en el seguimiento",
              font=("Segoe UI", 10, "bold")).pack(anchor="w")
    ttk.Label(tab_vis, foreground="#555", wraplength=640, justify="left",
              text="Copiá la columna POLIZA del archivo de seguimiento y pegala acá. "
                   "Si es el primer cruce del turno, dejalo vacío.").pack(anchor="w", pady=(2, 4))
    marco = ttk.Frame(tab_vis); marco.pack(fill="both", expand=True)
    scroll = ttk.Scrollbar(marco); scroll.pack(side="right", fill="y")
    txt = tk.Text(marco, height=7, yscrollcommand=scroll.set, font=("Consolas", 9))
    txt.pack(side="left", fill="both", expand=True)
    scroll.config(command=txt.yview)

    lbl_cont = ttk.Label(tab_vis, text="0 pólizas detectadas · modo PRIMER CRUCE",
                         foreground="#666")
    lbl_cont.pack(anchor="w", pady=(4, 8))

    def actualizar(_e=None):
        n = len(parsear_polizas_pegadas(txt.get("1.0", "end")))
        modo = "PRIMER CRUCE" if n == 0 else "CRUCE ACUMULATIVO"
        lbl_cont.config(text=f"{n} pólizas detectadas · modo {modo}")

    txt.bind("<KeyRelease>", actualizar)
    txt.bind("<<Paste>>", lambda e: root.after(50, actualizar))

    def ejecutar_vis():
        if not estado["bajada"]:
            messagebox.showwarning("Falta el archivo",
                                   "Seleccioná o importá la bajada del visualizador.")
            return
        elegidos = [n for n, v in estado["vars"].items() if v.get()]
        try:
            headers, filas, st, apartadas = procesar_cruce(
                estado["bajada"], txt.get("1.0", "end"), elegidos)
        except Exception as exc:
            messagebox.showerror("Error al procesar", str(exc))
            return

        if st["desconocidos"]:
            if not messagebox.askyesno("Auditores no reconocidos",
                    "Estos auditores aparecen pero NO están tildados:\n\n"
                    + "\n".join(st["desconocidos"])
                    + f"\n\nSe descartaron {st['filtrados']} filas.\n\n¿Continuar?"):
                return

        if st["nuevos"] == 0 and not st["apartadas"]:
            messagebox.showinfo("Sin hallazgos nuevos",
                f"Se revisaron {st['total_filas']} filas.\n"
                f"{st['total_con_obs']} con observación.\n"
                f"Ya cargadas: {st['ya_cargados']} · Filtradas: {st['filtrados']}\n\n"
                "No se generó ningún archivo.")
            return

        det = [f"Filas en la bajada: {st['total_filas']}",
               f"Con observación: {st['total_con_obs']}",
               f"Ya cargadas (descartadas): {st['ya_cargados']}"]
        if st["filtrados"]:
            det.append(f"Filtradas por verificador: {st['filtrados']}")
        if st["duplicados"]:
            det.append(f"Duplicados colapsados: {st['duplicados']}")
        det.append(f"Observaciones limpiadas: {st['limpiadas']}")
        if st["sin_categoria"]:
            det.append(f"Sin categoría reconocida: {st['sin_categoria']}")
        if st["multi_autor"]:
            det.append(f"ATENCIÓN · con más de un auditor: {st['multi_autor']}")
        if st["apartadas"]:
            det.append(f"\nAPARTADAS (hoja DESCONOCIDOS): {st['apartadas']}")
        det.append(f"\nHALLAZGOS NUEVOS: {st['nuevos']}")
        guardar_resultado(headers, filas, st["turno"], "Cruce", det,
                          apartadas, st["headers_origen"])

    ttk.Button(tab_vis, text="Ejecutar cruce", command=ejecutar_vis).pack(fill="x", ipady=6)

    # =====================================================================
    # PESTAÑA 2 — MUESTRA DE NATURGY
    # =====================================================================
    ttk.Label(tab_nat, text="Archivo de la muestra",
              font=("Segoe UI", 10, "bold")).pack(anchor="w")
    ttk.Label(tab_nat, foreground="#555", wraplength=640, justify="left",
              text="Seleccioná el archivo que envía Naturgy (RESULTADOS DE LA MUESTRA). "
                   "Se convierte al mismo formato del cruce, con ORIGEN = CLIENTE.\n\n"
                   "Entran todas las filas: las correctas y las incorrectas. "
                   "No se filtra ni se descarta nada.").pack(anchor="w", pady=(2, 10))

    fm = ttk.Frame(tab_nat); fm.pack(fill="x")
    lbl_muestra = ttk.Label(fm, text="Ningún archivo seleccionado", foreground="#888")

    def elegir_muestra():
        ruta = filedialog.askopenfilename(
            title="Seleccioná la muestra de Naturgy",
            filetypes=[("Excel", "*.xlsx *.xlsm *.XLSX"), ("Todos", "*.*")])
        if ruta:
            estado["muestra"] = ruta
            lbl_muestra.config(text=os.path.basename(ruta), foreground="#000")

    ttk.Button(fm, text="Seleccionar archivo...", command=elegir_muestra).pack(side="left")
    lbl_muestra.pack(side="left", padx=10)

    info = ttk.LabelFrame(tab_nat, text="Qué hace con los datos", padding=10)
    info.pack(fill="x", pady=(16, 0))
    for t in ["• Combina CALLE + FINCA en una sola columna CALLE.",
              "• Arma DATOS DOMICILIO con PORTAL, ESCALERA, PISO y PUERTA.",
              "• Reconstruye la fecha y la hora (30826 → 03/08/26, 901 → 09:01).",
              "• Pasa CATEGORIA a mayúsculas.",
              "• Unifica los tipos numéricos para que convivan con el cruce."]:
        ttk.Label(info, text=t, foreground="#444", wraplength=600,
                  justify="left").pack(anchor="w", pady=1)

    def ejecutar_nat():
        if not estado["muestra"]:
            messagebox.showwarning("Falta el archivo", "Seleccioná la muestra de Naturgy.")
            return
        try:
            headers, filas, st = procesar_muestra(estado["muestra"])
        except Exception as exc:
            messagebox.showerror("Error al procesar", str(exc))
            return
        if st["columnas_ausentes"]:
            if not messagebox.askyesno("Columnas faltantes",
                    "La muestra no trae estas columnas y saldrán vacías:\n\n"
                    + "\n".join(st["columnas_ausentes"]) + "\n\n¿Continuar?"):
                return
        if st["nuevos"] == 0:
            messagebox.showinfo("Sin filas", "La muestra no tiene filas de datos.")
            return
        det = [f"Filas procesadas: {st['total_filas']}"]
        if st["sin_fecha"]:
            det.append(f"Sin fecha reconocible: {st['sin_fecha']}")
        if st["sin_hora"]:
            det.append(f"Sin hora reconocible: {st['sin_hora']}")
        if st["categorias"]:
            det.append("\nCategorías:")
            for k, n in sorted(st["categorias"].items(), key=lambda x: -x[1]):
                det.append(f"   {n:>4}  {k}")
        det.append(f"\nFILAS EN EL ARCHIVO: {st['nuevos']}")
        guardar_resultado(headers, filas, st["turno"], "Muestra", det)

    ttk.Button(tab_nat, text="Procesar muestra", command=ejecutar_nat).pack(
        fill="x", ipady=6, pady=(16, 0))

    # =====================================================================
    # PIE
    # =====================================================================
    def abrir_config_errores():
        simples, codigos = cargar_errores()
        win = tk.Toplevel(root); win.title("Errores conocidos")
        win.geometry("520x520"); win.transient(root); win.grab_set()
        c = ttk.Frame(win, padding=14); c.pack(fill="both", expand=True)
        ttk.Label(c, text="Errores conocidos", font=("Segoe UI", 12, "bold")).pack(anchor="w")
        ttk.Label(c, foreground="#555", wraplength=470, justify="left",
                  text="Sirven para deducir la CATEGORIA cuando el visualizador no la "
                       "trae cargada. Esta lista es compartida por todos.").pack(
            anchor="w", pady=(2, 10))
        ttk.Label(c, text="Con código de anomalía detrás (G, K, 34, ZT...)",
                  font=("Segoe UI", 9, "bold")).pack(anchor="w")
        lb_cod = tk.Listbox(c, height=3, font=("Consolas", 9))
        lb_cod.pack(fill="x", pady=(2, 8))
        for e in codigos:
            lb_cod.insert("end", e)
        ttk.Label(c, text="Errores literales", font=("Segoe UI", 9, "bold")).pack(anchor="w")
        lb_sim = tk.Listbox(c, height=9, font=("Consolas", 9))
        lb_sim.pack(fill="both", expand=True, pady=(2, 8))
        for e in simples:
            lb_sim.insert("end", e)

        def agregar_a(lb):
            t = simpledialog.askstring("Agregar error", "Texto exacto del error:", parent=win)
            if not t or not t.strip():
                return
            v = t.strip()
            if v.upper() in [lb.get(i).upper() for i in range(lb.size())]:
                messagebox.showinfo("Ya existe", "Ese error ya está.", parent=win)
                return
            lb.insert("end", v)

        def seleccionada():
            if lb_sim.curselection():
                return lb_sim
            if lb_cod.curselection():
                return lb_cod
            return None

        def modificar():
            lb = seleccionada()
            if lb is None:
                messagebox.showinfo("Modificar", "Seleccioná un error.", parent=win)
                return
            i = lb.curselection()[0]
            t = simpledialog.askstring("Modificar error", "Texto exacto del error:",
                                       initialvalue=lb.get(i), parent=win)
            if not t or not t.strip():
                return
            lb.delete(i); lb.insert(i, t.strip()); lb.selection_set(i)

        def quitar_e():
            lb = seleccionada()
            if lb is None:
                messagebox.showinfo("Quitar", "Seleccioná un error.", parent=win)
                return
            for i in reversed(list(lb.curselection())):
                lb.delete(i)

        fb = ttk.Frame(c); fb.pack(fill="x", pady=(0, 8))
        ttk.Button(fb, text="Agregar literal", command=lambda: agregar_a(lb_sim)).pack(side="left")
        ttk.Button(fb, text="Agregar con código",
                   command=lambda: agregar_a(lb_cod)).pack(side="left", padx=6)
        ttk.Button(fb, text="Modificar", command=modificar).pack(side="left")
        ttk.Button(fb, text="Quitar", command=quitar_e).pack(side="left", padx=6)
        lb_sim.bind("<<ListboxSelect>>",
                    lambda e: lb_cod.selection_clear(0, "end") if lb_sim.curselection() else None)
        lb_cod.bind("<<ListboxSelect>>",
                    lambda e: lb_sim.selection_clear(0, "end") if lb_cod.curselection() else None)

        def guardar_y_cerrar():
            s = [lb_sim.get(i) for i in range(lb_sim.size())]
            k = [lb_cod.get(i) for i in range(lb_cod.size())]
            if not guardar_errores(s, k):
                messagebox.showwarning("Sin permisos",
                                       "Vale para esta sesión, pero no se guardó.", parent=win)
            win.destroy()

        ttk.Button(c, text="Guardar y cerrar", command=guardar_y_cerrar).pack(fill="x", ipady=4)

    pie = ttk.Frame(cont); pie.pack(fill="x", pady=(8, 0))
    ttk.Button(pie, text="Configurar errores conocidos...",
               command=abrir_config_errores).pack(side="left")

    _rc = ruta_config()
    _compartida = not _rc.startswith(_carpeta_usuario())
    if not config_disponible() and os.path.isfile(ruta_cache_config()):
        _txt = f"v{VERSION} · SIN CONEXION a la config compartida (usando la última conocida)"
        _col = "#B00"
    else:
        _txt = (f"v{VERSION} · configuración "
                f"{'COMPARTIDA' if _compartida else 'INDIVIDUAL'}: {_rc}")
        _col = "#999"
    ttk.Label(cont, text=_txt, foreground=_col,
              font=("Segoe UI", 8)).pack(anchor="w", pady=(6, 0))

    root.mainloop()


if __name__ == "__main__":
    lanzar_ui()
